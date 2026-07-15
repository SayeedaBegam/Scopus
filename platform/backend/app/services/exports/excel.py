import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session
from app.models import Department,Institution,InternationalCollaboration,Lab,Professor,ProfessorPublication,Publication,ReviewItem,SyncRun

def safe(value):
    if isinstance(value,str) and value[:1] in ("=","+","-","@"): return "'"+value
    return value
def sheet(wb,name,headers,rows):
    ws=wb.create_sheet(name[:31]); ws.append(headers)
    for cell in ws[1]: cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="005A70")
    for row in rows: ws.append([safe(x) for x in row])
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=min(55,max(12,max(len(str(x.value or "")) for x in col)+2))

def export_workbook(db:Session,professor_id:int|None=None):
    wb=Workbook(); wb.remove(wb.active)
    professors=db.scalars(select(Professor).where(Professor.id==professor_id) if professor_id else select(Professor)).all(); pids=[p.id for p in professors]
    links=db.scalars(select(ProfessorPublication).where(ProfessorPublication.professor_id.in_(pids))).all() if pids else []; pubids={x.publication_id for x in links}
    pubs=db.scalars(select(Publication).where(Publication.id.in_(pubids))).all() if pubids else []; collabs=db.scalars(select(InternationalCollaboration).where(InternationalCollaboration.professor_id.in_(pids))).all() if pids else []
    insts={x.id:x for x in db.scalars(select(Institution)).all()}; profs={x.id:x for x in professors}; pubmap={x.id:x for x in pubs}
    sheet(wb,"Summary",["Metric","Value"],[["Generated",datetime.utcnow().isoformat()], ["Professors",len(professors)],["Publications",len(pubs)],["International collaborations",len(collabs)],["Countries",len({x.country for x in collabs})],["Institutions",len({x.institution_id for x in collabs})]])
    sheet(wb,"International Collaborations",["Year","Number of Authors","Source Title","Publication Title","UTN Researcher","International Researcher","International Institution","Partner Department or Lab","Country","DOI","Scopus EID","Review Status","Notes"],[[c.year,pubmap[c.publication_id].author_count,pubmap[c.publication_id].source_title,pubmap[c.publication_id].title,profs[c.professor_id].full_name,c.international_coauthor,insts[c.institution_id].canonical_name,c.partner_department,c.country,pubmap[c.publication_id].doi,pubmap[c.publication_id].scopus_eid,c.review_status.value,c.notes] for c in collabs])
    sheet(wb,"Publications",["Year","Title","Source","Type","DOI","Scopus EID","Authors","Citations"],[[p.year,p.title,p.source_title,p.document_type,p.doi,p.scopus_eid,p.author_count,p.citation_count] for p in pubs])
    sheet(wb,"Professors",["Name","Email","ORCID","Scopus Author ID","Active"],[[p.full_name,p.email,p.orcid,p.scopus_author_id,p.is_active] for p in professors])
    sheet(wb,"Partner Institutions",["Institution","Country","Verified"],[[i.canonical_name,i.country,i.verified] for i in {insts[c.institution_id] for c in collabs}])
    sheet(wb,"Partner Countries",["Country","Collaborations"],[[country,sum(x.country==country for x in collabs)] for country in sorted({x.country for x in collabs})])
    sheet(wb,"Data Quality Notes",["Entity","Reason","Status","Notes"],[[r.entity_type,r.reason,r.status.value,r.notes] for r in db.scalars(select(ReviewItem)).all()])
    if professor_id is None: sheet(wb,"Synchronization History",["Professor ID","Type","Status","Fetched","Created","Updated","Error"],[[s.professor_id,s.sync_type,s.status,s.records_fetched,s.records_created,s.records_updated,s.error_message] for s in db.scalars(select(SyncRun)).all()])
    data=io.BytesIO(); wb.save(data); return data.getvalue()
