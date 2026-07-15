"""
Excel export utilities for generating professional publication overview sheets.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from typing import Optional
import io


def create_professional_excel_export(df: pd.DataFrame, professor_name: Optional[str] = None) -> bytes:
    """
    Create a professional Excel sheet with international collaboration data.
    
    Args:
        df: Processed DataFrame with collaboration data
        professor_name: Optional filter for specific professor
        
    Returns:
        Excel file bytes for download
    """
    
    if professor_name:
        df = df[df['Professor Name'] == professor_name].copy()
        title = f"International Collaborations - {professor_name}"
        filename = f"collaborations_{professor_name.replace(' ', '_')}.xlsx"
    else:
        title = "International Joint Publications - Overview"
        filename = "international_collaborations.xlsx"
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Collaborations"
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Add title
    ws['A1'] = title
    ws['A1'].font = title_font
    ws.merge_cells('A1:K1')
    ws['A1'].alignment = center_alignment
    
    # Add metadata
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = Font(italic=True, size=9, color="666666")
    
    # Add summary statistics
    row = 4
    ws[f'A{row}'] = "Summary Statistics"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    
    row += 1
    ws[f'A{row}'] = f"Total Collaborations:"
    ws[f'B{row}'] = len(df)
    
    row += 1
    ws[f'A{row}'] = f"Partner Countries:"
    ws[f'B{row}'] = df['Country'].nunique()
    
    row += 1
    ws[f'A{row}'] = f"Partner Institutions:"
    ws[f'B{row}'] = df['International Partner Institution'].nunique()
    
    row += 1
    ws[f'A{row}'] = f"Professors:"
    ws[f'B{row}'] = df['Professor Name'].nunique()
    
    # Add empty row
    row += 2
    
    # Add data header
    headers = [
        'Professor Name',
        'Year',
        'Publication Title',
        'Source Title',
        'Authors',
        'Partner Institution',
        'Country',
        'DOI',
        'Needs Review',
        'Notes'
    ]
    
    header_row = row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Add data rows
    row += 1
    for idx, record in df.iterrows():
        ws.cell(row=row, column=1).value = record.get('Professor Name', '')
        ws.cell(row=row, column=2).value = record.get('Year', '')
        ws.cell(row=row, column=3).value = record.get('Publication Title', '')
        ws.cell(row=row, column=4).value = record.get('Source Title', '')
        ws.cell(row=row, column=5).value = record.get('Authors', '')
        ws.cell(row=row, column=6).value = record.get('International Partner Institution', '')
        ws.cell(row=row, column=7).value = record.get('Country', '')
        ws.cell(row=row, column=8).value = record.get('DOI', '')
        ws.cell(row=row, column=9).value = record.get('Needs Review', '')
        ws.cell(row=row, column=10).value = record.get('Notes', '')
        
        # Apply formatting to data row
        for col_num in range(1, 11):
            cell = ws.cell(row=row, column=col_num)
            cell.border = border
            if col_num in [1, 7, 9]:  # Professor, Country, Needs Review
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 20
    
    # Freeze header
    ws.freeze_panes = f'A{header_row + 1}'
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def create_country_summary_sheet(df: pd.DataFrame) -> bytes:
    """Create a summary sheet organized by country."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "By Country"
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    country_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    country_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    ws['A1'] = "International Collaborations by Country"
    ws['A1'].font = Font(bold=True, size=14)
    
    row = 3
    
    # Get countries
    countries = sorted(df['Country'].unique())
    
    for country in countries:
        country_data = df[df['Country'] == country]
        
        # Country header
        ws[f'A{row}'] = f"{country} ({len(country_data)} collaborations)"
        ws[f'A{row}'].font = country_font
        ws[f'A{row}'].fill = country_fill
        ws.merge_cells(f'A{row}:G{row}')
        
        row += 1
        
        # Column headers for country section
        headers = ['Institution', 'Professor', 'Year', 'Publication Title', 'Source', 'Needs Review', 'Notes']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        row += 1
        
        # Data for this country
        for idx, record in country_data.iterrows():
            ws.cell(row=row, column=1).value = record['International Partner Institution']
            ws.cell(row=row, column=2).value = record['Professor Name']
            ws.cell(row=row, column=3).value = record['Year']
            ws.cell(row=row, column=4).value = record['Publication Title']
            ws.cell(row=row, column=5).value = record['Source Title']
            ws.cell(row=row, column=6).value = record['Needs Review']
            ws.cell(row=row, column=7).value = record['Notes']
            
            for col_num in range(1, 8):
                cell = ws.cell(row=row, column=col_num)
                cell.border = border
            
            row += 1
        
        row += 1  # Empty row between countries
    
    # Adjust columns
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def create_manual_tracking_excel(df: pd.DataFrame, professor_name: Optional[str] = None) -> bytes:
    """Create an Excel sheet for the cleaned manual CSV workflow."""
    working_df = df.copy()

    if professor_name:
        working_df = working_df[working_df["UTN Researcher (s)"] == professor_name].copy()
        title = f"International Joint Publications - {professor_name}"
    else:
        title = "International Joint Publications - Overview"

    wb = Workbook()
    ws = wb.active
    ws.title = "Joint Publications"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws["A1"] = title
    ws["A1"].font = title_font
    ws.merge_cells("A1:K1")
    ws["A1"].alignment = center_alignment

    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, size=9, color="666666")

    row = 4
    ws[f"A{row}"] = "Summary Statistics"
    ws[f"A{row}"].font = subheader_font
    ws[f"A{row}"].fill = subheader_fill

    summary_rows = [
        ("Total Rows", len(working_df)),
        ("Professors", working_df["UTN Researcher (s)"].nunique()),
        ("Partner Institutions", working_df["Other University/Institution"].nunique()),
        ("Countries", working_df["Country"].nunique()),
    ]

    for label, value in summary_rows:
        row += 1
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value

    row += 2
    headers = list(working_df.columns)
    header_row = row

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    row += 1
    for _, record in working_df.iterrows():
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = record.get(header, "")
            cell.border = border
            cell.alignment = center_alignment if header in {"Year", "No of Author(s)", "Country"} else left_alignment
        row += 1

    column_widths = {
        "A": 10,
        "B": 14,
        "C": 28,
        "D": 42,
        "E": 22,
        "F": 20,
        "G": 22,
        "H": 28,
        "I": 32,
        "J": 28,
        "K": 16,
    }
    for column_letter, width in column_widths.items():
        ws.column_dimensions[column_letter].width = width

    ws.freeze_panes = f"A{header_row + 1}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
