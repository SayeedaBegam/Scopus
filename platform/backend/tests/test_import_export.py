from io import BytesIO
from openpyxl import load_workbook

def test_semicolon_csv_mapping(admin):
    data=b'Authors;Document Title;Year;Source Title;Affiliations;DOI\n"A, B";Test;2026;Journal;"UTN, Germany | Oxford, United Kingdom";10.1/x\n'
    r=admin.post("/api/v1/imports/scopus-csv",files={"file":("sample.csv",data,"text/csv")})
    assert r.status_code==200 and not r.json()["missing_required"]

def test_excel_has_expected_sheets(admin):
    r=admin.post("/api/v1/exports/overall")
    assert r.status_code==200
    wb=load_workbook(BytesIO(r.content))
    assert {"Summary","International Collaborations","Publications","Professors","Data Quality Notes"}.issubset(wb.sheetnames)
